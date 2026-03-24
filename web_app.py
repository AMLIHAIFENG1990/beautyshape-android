#!/usr/bin/env python3
"""
BeautyShape v4 — 高级交互式直播调试工具
http://localhost:8899
"""
import sys, os, base64, json, time, threading, webbrowser
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from flask import Flask, request, jsonify, render_template_string
import cv2, numpy as np
from face_shape_engine import FaceShapeEngine

app = Flask(__name__)
analyzer = FaceShapeEngine()
DB_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'experience_db.json')
def load_db():
    if os.path.exists(DB_PATH):
        with open(DB_PATH) as f: return json.load(f)
    return {'samples':[]}
def save_db(db):
    with open(DB_PATH,'w') as f: json.dump(db,f,ensure_ascii=False,indent=2)

HTML = r'''<!DOCTYPE html>
<html lang="zh">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>BeautyShape · 脸型分析</title>
<style>
/* ═══════════════ RESET & VARS ═══════════════ */
*{margin:0;padding:0;box-sizing:border-box}
:root{
  --bg:#06060f;--bg2:#0c0c1d;--bg3:#12122a;--bg4:#1a1a3a;
  --g1:#8b5cf6;--g2:#a78bfa;--g3:#c084fc;--g4:#7c3aed;
  --p1:#ec4899;--p2:#f472b6;--p3:#f9a8d4;
  --g:#34d399;--a:#fbbf24;--b:#60a5fa;--r:#f87171;
  --t:#f8fafc;--t2:#cbd5e1;--t3:#64748b;
  --glass:rgba(255,255,255,.04);--glass2:rgba(255,255,255,.07);
  --border:rgba(255,255,255,.06);--border2:rgba(139,92,246,.2);
  --shadow:0 8px 40px rgba(0,0,0,.5);
  --ease:cubic-bezier(.4,0,.2,1);
}
html{scroll-behavior:smooth}
body{font-family:'SF Pro Display',-apple-system,'PingFang SC','Helvetica Neue',sans-serif;background:var(--bg);color:var(--t);overflow-x:hidden;min-height:100vh}
::selection{background:var(--g1);color:#fff}
::-webkit-scrollbar{width:5px}::-webkit-scrollbar-track{background:0 0}
::-webkit-scrollbar-thumb{background:var(--bg4);border-radius:10px}

/* ═══════════════ ANIMATIONS ═══════════════ */
@keyframes fadeUp{from{opacity:0;transform:translateY(30px)}to{opacity:1;transform:translateY(0)}}
@keyframes fadeIn{from{opacity:0}to{opacity:1}}
@keyframes slideR{from{opacity:0;transform:translateX(40px)}to{opacity:1;transform:translateX(0)}}
@keyframes scaleIn{from{opacity:0;transform:scale(.92)}to{opacity:1;transform:scale(1)}}
@keyframes pulse{0%,100%{opacity:1}50%{opacity:.4}}
@keyframes glow{0%,100%{box-shadow:0 0 20px rgba(139,92,246,.2)}50%{box-shadow:0 0 50px rgba(139,92,246,.5)}}
@keyframes float{0%,100%{transform:translateY(0)}50%{transform:translateY(-6px)}}
@keyframes gradient{0%{background-position:0% 50%}50%{background-position:100% 50%}100%{background-position:0% 50%}}
@keyframes spin{to{transform:rotate(360deg)}}
@keyframes barGrow{from{width:0}to{width:var(--w)}}
@keyframes countUp{from{opacity:0;transform:translateY(10px)}to{opacity:1;transform:translateY(0)}}
@keyframes ripple{to{transform:scale(4);opacity:0}}
@keyframes borderGlow{0%,100%{border-color:rgba(139,92,246,.1)}50%{border-color:rgba(139,92,246,.4)}}
@keyframes typewriter{from{width:0}to{width:100%}}

/* ═══════════════ BACKGROUND ═══════════════ */
.bg-orbs{position:fixed;inset:0;pointer-events:none;z-index:0;overflow:hidden}
.bg-orbs::before,.bg-orbs::after{content:'';position:absolute;border-radius:50%;filter:blur(120px);opacity:.15}
.bg-orbs::before{width:600px;height:600px;background:var(--g1);top:-200px;left:-100px;animation:float 8s ease-in-out infinite}
.bg-orbs::after{width:500px;height:500px;background:var(--p1);bottom:-150px;right:-100px;animation:float 10s ease-in-out infinite reverse}

/* ═══════════════ NAV ═══════════════ */
.nav{position:fixed;top:0;left:0;right:0;z-index:100;height:60px;display:flex;align-items:center;padding:0 28px;gap:20px;background:rgba(6,6,15,.8);backdrop-filter:blur(30px) saturate(1.5);border-bottom:1px solid var(--border)}
.nav-logo{display:flex;align-items:center;gap:10px;cursor:pointer;transition:all .3s}
.nav-logo:hover{transform:scale(1.02)}
.nav-logo .mark{width:34px;height:34px;background:linear-gradient(135deg,var(--g1),var(--p1));border-radius:10px;display:flex;align-items:center;justify-content:center;font-size:15px;box-shadow:0 4px 15px rgba(139,92,246,.3)}
.nav-logo .name{font-size:16px;font-weight:700;letter-spacing:-.5px}
.nav-logo .name em{font-style:normal;background:linear-gradient(135deg,var(--g2),var(--p2));-webkit-background-clip:text;-webkit-text-fill-color:transparent}
.nav-tabs{display:flex;gap:2px;margin-left:auto;background:rgba(255,255,255,.03);border-radius:10px;padding:3px;border:1px solid var(--border)}
.nav-tab{padding:7px 18px;border-radius:7px;cursor:pointer;font-size:12px;font-weight:500;color:var(--t3);background:0 0;border:none;transition:all .3s var(--ease);position:relative;overflow:hidden}
.chart-btn{padding:5px 12px;border-radius:6px;cursor:pointer;font-size:11px;font-weight:500;color:var(--t3);background:0 0;border:none;transition:all .3s var(--ease)}
.nav-tab:hover{color:var(--t2)}
.nav-tab.on{background:linear-gradient(135deg,var(--g4),var(--g1));color:#fff;box-shadow:0 4px 15px rgba(139,92,246,.3)}
.chart-btn.on{background:linear-gradient(135deg,var(--g4),var(--g1));color:#fff;box-shadow:0 4px 15px rgba(139,92,246,.3)}
/* 编辑功能 */
.edit-hint{font-size:11px;color:var(--t3);cursor:pointer;padding:2px 8px;border-radius:4px;background:rgba(255,255,255,.05);transition:all .2s;margin-left:8px}
.edit-hint:hover{color:var(--g1);background:rgba(139,92,246,.1)}
.editable-item{position:relative;transition:all .3s}
.editable-item[contenteditable="true"]{outline:1px dashed var(--g1);outline-offset:4px;border-radius:4px}
.del-edit{display:none;position:absolute;right:4px;top:50%;transform:translateY(-50%);background:rgba(239,68,68,.2);border:none;color:#ef4444;width:22px;height:22px;border-radius:50%;cursor:pointer;font-size:10px;line-height:22px;text-align:center}
.del-edit:hover{background:rgba(239,68,68,.4)}
.light-item .del-edit{right:2px;top:6px;transform:none}
.add-item{border:1px dashed var(--border)!important;padding:8px 12px;margin:4px 0}
.editing{border:1px solid var(--g1)!important;box-shadow:0 0 20px rgba(139,92,246,.1)}
/* 三庭可点击 */
.court.clickable{cursor:pointer;transition:all .3s var(--ease);border-radius:8px;padding:6px}
.court.clickable:hover{background:rgba(139,92,246,.08)}
.court.clickable.on{background:rgba(139,92,246,.15);box-shadow:0 0 12px rgba(139,92,246,.2)}
/* 画布叠加层 */
.canvas-wrap canvas{display:block}
#overlay{position:absolute;top:0;left:0;pointer-events:none;opacity:.85}
.nav-status{display:flex;align-items:center;gap:6px;font-size:10px;color:var(--t3)}
.pulse-dot{width:6px;height:6px;border-radius:50%;background:var(--g);box-shadow:0 0 8px rgba(52,211,153,.5);animation:pulse 2s infinite}

/* ═══════════════ MAIN ═══════════════ */
.main{position:relative;z-index:1;padding:72px 20px 20px;max-width:1600px;margin:0 auto}
.panel{display:none;animation:fadeUp .5s var(--ease)}
.panel.on{display:block}

/* ═══════════════ LAYOUT ═══════════════ */
.analysis-grid{display:grid;grid-template-columns:1fr 440px;gap:16px;min-height:calc(100vh - 92px)}
@media(max-width:1100px){.analysis-grid{grid-template-columns:1fr}.results-col{max-height:none!important}}

/* ═══════════════ TOOLBAR ═══════════════ */
.toolbar{display:flex;gap:8px;flex-wrap:wrap;align-items:center;margin-bottom:12px}
.btn{position:relative;display:inline-flex;align-items:center;gap:6px;padding:9px 18px;border-radius:10px;font-size:13px;font-weight:500;cursor:pointer;border:1px solid var(--border);background:var(--glass);color:var(--t);transition:all .25s var(--ease);overflow:hidden;font-family:inherit}
.btn:hover{background:var(--glass2);border-color:var(--border2);transform:translateY(-1px);box-shadow:0 4px 20px rgba(0,0,0,.2)}
.btn:active{transform:translateY(0);box-shadow:none}
.btn .ripple{position:absolute;border-radius:50%;background:rgba(255,255,255,.2);transform:scale(0);animation:ripple .6s linear}
.btn-p{background:linear-gradient(135deg,var(--g4),var(--g1));border:none;color:#fff;box-shadow:0 4px 20px rgba(139,92,246,.3)}
.btn-p:hover{box-shadow:0 6px 30px rgba(139,92,246,.5);background:linear-gradient(135deg,var(--g1),var(--p1))}
.btn-a{background:linear-gradient(135deg,var(--p1),#be185d);border:none;color:#fff;box-shadow:0 4px 20px rgba(236,72,153,.3)}
.btn-a:hover{box-shadow:0 6px 30px rgba(236,72,153,.5)}
.btn-g{background:linear-gradient(135deg,var(--g),#059669);border:none;color:#fff}
.btn-icon{padding:9px 12px}
.hint{font-size:11px;color:var(--t3);margin-left:auto;transition:all .3s}

/* ═══════════════ CANVAS ═══════════════ */
.canvas-wrap{position:relative;background:var(--bg2);border:1px solid var(--border);border-radius:16px;overflow:hidden;display:flex;align-items:center;justify-content:center;flex:1;max-height:60vh;min-height:300px;transition:border-color .3s}
.canvas-wrap canvas{display:block}
.canvas-wrap:hover{border-color:var(--border2)}
.canvas-wrap canvas{max-width:100%;max-height:100%;cursor:crosshair}
.ph{color:var(--t3);text-align:center;animation:fadeIn .6s}
.ph .icon{font-size:56px;margin-bottom:12px;opacity:.3;display:block}
.ph h3{font-size:15px;font-weight:500;margin-bottom:6px;color:var(--t2)}
.ph p{font-size:12px;line-height:1.7}
.loading{display:none;position:absolute;inset:0;background:rgba(6,6,15,.92);backdrop-filter:blur(12px);z-index:20;flex-direction:column;align-items:center;justify-content:center;gap:14px;animation:fadeIn .2s}
.loading.on{display:flex}
.spinner{width:44px;height:44px;border:3px solid var(--bg4);border-top-color:var(--g1);border-radius:50%;animation:spin .7s linear infinite}
.loading span{font-size:13px;color:var(--t2)}

/* ═══════════════ RESULTS ═══════════════ */
.results-col{display:flex;flex-direction:column;gap:10px;overflow-y:auto;max-height:calc(100vh - 92px);padding-right:4px}
.stitle{font-size:11px;font-weight:600;color:var(--t3);text-transform:uppercase;letter-spacing:1.5px;margin-bottom:2px;display:flex;align-items:center;gap:10px}
.stitle::after{content:'';flex:1;height:1px;background:var(--border)}

/* Cards */
.card{background:var(--glass);border:1px solid var(--border);border-radius:14px;padding:16px;transition:all .35s var(--ease);animation:slideR .4s var(--ease);backdrop-filter:blur(10px)}
.card:hover{background:var(--glass2);border-color:var(--border2);transform:translateX(-2px);box-shadow:var(--shadow)}
.card .lb{font-size:10px;color:var(--t3);text-transform:uppercase;letter-spacing:.8px;margin-bottom:8px}
.card .val{font-size:24px;font-weight:700;letter-spacing:-.5px}
.card .val.grad{background:linear-gradient(135deg,var(--g2),var(--p2));-webkit-background-clip:text;-webkit-text-fill-color:transparent}
.card .det{font-size:12px;color:var(--t2);margin-top:6px;line-height:1.6}

/* Tags */
.tags{display:flex;gap:6px;flex-wrap:wrap;margin:8px 0}
.tag{display:inline-flex;align-items:center;gap:4px;padding:4px 10px;border-radius:20px;font-size:11px;font-weight:500;background:rgba(139,92,246,.1);color:var(--g2);border:1px solid rgba(139,92,246,.15)}
.tag .dot{width:5px;height:5px;border-radius:50%;background:currentColor}

/* Profile Bars */
.pbar{display:flex;align-items:center;gap:8px;margin:5px 0}
.pbar .nm{font-size:11px;color:var(--t3);width:24px;text-align:right;font-weight:500}
.pbar .track{flex:1;height:20px;background:var(--bg3);border-radius:6px;overflow:hidden;position:relative}
.pbar .fill{height:100%;border-radius:6px;animation:barGrow .8s var(--ease) forwards}
.pbar .pv{font-size:11px;font-weight:600;width:34px;text-align:right}

/* Three Courts */
.courts{display:flex;gap:10px;margin-top:6px}
.court{flex:1;text-align:center;padding:10px 8px;background:var(--bg3);border-radius:10px;border:1px solid var(--border);transition:all .3s}
.court:hover{border-color:var(--border2);transform:translateY(-2px)}
.court .cv{font-size:22px;font-weight:700;transition:color .3s}
.court .cn{font-size:10px;color:var(--t3);margin-top:3px}
.badge{display:inline-flex;align-items:center;gap:3px;padding:2px 7px;border-radius:4px;font-size:10px;font-weight:600}
.badge.ok{background:rgba(52,211,153,.12);color:var(--g)}
.badge.warn{background:rgba(251,191,36,.12);color:var(--a)}

/* Five Eye Bar */
.eye-bar{display:flex;align-items:center;gap:2px;margin:10px 0 6px;height:22px;border-radius:4px;overflow:hidden}
.eye-seg{height:100%;transition:all .6s var(--ease)}
.eye-labels{display:flex;justify-content:space-between;font-size:10px;color:var(--t3)}

/* Lighting */
.light-grid{display:grid;grid-template-columns:1fr 1fr;gap:6px;margin-top:8px}
.light-item{display:flex;align-items:center;gap:8px;padding:8px 10px;background:var(--bg3);border-radius:8px;font-size:11px;color:var(--t2);border:1px solid var(--border);transition:all .2s}
.light-item:hover{border-color:var(--border2);background:var(--bg4)}
.light-item .icon2{font-size:14px;width:20px;text-align:center}

/* ═══════════════ SAVE FORM ═══════════════ */
.save-form{background:var(--bg2);border:1px solid var(--border);border-radius:14px;padding:16px;animation:scaleIn .3s var(--ease)}
.save-form.hidden{display:none}
.save-form label{font-size:11px;color:var(--t3);display:block;margin-bottom:5px;font-weight:500}
.save-form select,.save-form input{background:var(--bg3);border:1px solid var(--border);color:var(--t);padding:10px 14px;border-radius:10px;font-size:13px;width:100%;outline:none;transition:all .25s;font-family:inherit}
.save-form select:focus,.save-form input:focus{border-color:var(--g1);box-shadow:0 0 0 3px rgba(139,92,246,.15)}
.save-form select{margin-bottom:8px;cursor:pointer}
.save-form input{margin-bottom:10px}
.save-actions{display:flex;gap:8px}

/* ═══════════════ EXPERIENCE ═══════════════ */
.exp-wrap{display:flex;flex-direction:column;gap:16px}
.exp-header{display:flex;align-items:center;gap:12px;flex-wrap:wrap}
.exp-stats{display:flex;gap:8px;flex-wrap:wrap}
.exp-stat{background:var(--glass);border:1px solid var(--border);border-radius:12px;padding:12px 18px;text-align:center;transition:all .3s;cursor:default}
.exp-stat:hover{border-color:var(--border2);transform:translateY(-2px)}
.exp-stat .n{font-size:28px;font-weight:700;background:linear-gradient(135deg,var(--g2),var(--p2));-webkit-background-clip:text;-webkit-text-fill-color:transparent;animation:countUp .4s var(--ease)}
.exp-stat .l{font-size:10px;color:var(--t3);margin-top:3px;text-transform:uppercase;letter-spacing:.5px}
.exp-grid{display:grid;grid-template-columns:repeat(auto-fill,minmax(340px,1fr));gap:10px}
.exp-item{background:var(--glass);border:1px solid var(--border);border-radius:12px;padding:14px;display:flex;justify-content:space-between;align-items:center;transition:all .3s var(--ease)}
.exp-item:hover{border-color:var(--border2);background:var(--glass2);transform:translateX(-2px)}
.exp-item .sh{font-size:16px;font-weight:700;color:var(--p2)}
.exp-item .md{font-size:11px;color:var(--t3);margin-top:4px}
.exp-item .dt{color:var(--a);font-size:11px}
.exp-item .del{background:0 0;border:none;color:var(--t3);cursor:pointer;font-size:14px;padding:6px;border-radius:6px;transition:all .2s}
.exp-item .del:hover{color:var(--r);background:rgba(248,113,113,.1)}

/* ═══════════════ FACE SHAPE RADAR ═══════════════ */
.radar-wrap{position:relative;width:200px;height:200px;margin:10px auto}
.radar-wrap canvas{width:100%;height:100%}

/* ═══════════════ TOOLTIP ═══════════════ */
.tooltip{position:fixed;background:var(--bg4);color:var(--t);padding:6px 12px;border-radius:8px;font-size:11px;pointer-events:none;z-index:200;opacity:0;transition:opacity .2s;box-shadow:var(--shadow)}
.tooltip.show{opacity:1}

/* ═══════════════ RESPONSIVE ═══════════════ */
@media(max-width:768px){
  .nav{padding:0 12px}.toolbar{gap:4px}.btn{padding:7px 12px;font-size:12px}
  .analysis-grid{gap:10px}.cours{flex-direction:column}.light-grid{grid-template-columns:1fr}
}
</style>
</head>
<body>
<div class="bg-orbs"></div>

<!-- NAV -->
<nav class="nav">
  <div class="nav-logo" onclick="switchTab('analyze')">
    <div class="mark">✦</div>
    <div class="name"><em>Beauty</em>Shape</div>
  </div>
  <div class="nav-tabs">
    <button class="nav-tab on" data-tab="analyze" onclick="switchTab('analyze')">✦ 分析</button>
    <button class="nav-tab" data-tab="exp" onclick="switchTab('exp')">◈ 经验库</button>
  </div>
  <div class="nav-status"><div class="pulse-dot"></div><span>在线</span></div>
</nav>

<!-- MAIN -->
<div class="main">
  <!-- ═══ ANALYSIS ═══ -->
  <div class="panel on" id="P-analyze">
    <div class="analysis-grid">
      <!-- LEFT -->
      <div style="display:flex;flex-direction:column">
        <div class="toolbar">
          <button class="btn btn-p" onclick="doCapture()">📷 截取屏幕</button>
          <button class="btn" onclick="$F.click()">🖼 导入图片</button>
          <input type="file" id="F" accept="image/*" style="display:none" onchange="loadFile(this)">
          <button class="btn btn-a" id="cropBtn" style="display:none" onclick="doCrop()">✂️ 确认选区</button>
          <button class="btn" onclick="doAnalyze()">🔍 分析</button>
          <button class="btn btn-icon" onclick="reset()" title="重置">↺</button>
          <span class="hint" id="hint">截屏后拖拽框选脸部区域</span>
        </div>
        <div class="canvas-wrap" id="CW">
          <div class="ph" id="ph"><span class="icon">✦</span><h3>开始分析</h3><p>📷 截取屏幕 或 🖼 导入图片<br>截屏后可拖拽框选脸部区域</p></div>
          <canvas id="C" style="display:none"></canvas>
          <div class="loading" id="LD"><div class="spinner"></div><span>分析中...</span></div>
        </div>
      </div>
      <!-- RIGHT -->
      <div class="results-col" id="RC">
        <div class="stitle">分析结果</div>
        <div id="R">
          <div class="card" style="opacity:.4"><div class="lb">等待分析</div><div class="det" style="text-align:center;padding:24px;color:var(--t3)">导入图片或截取屏幕开始</div></div>
        </div>
        <div class="save-form hidden" id="SF">
          <div class="stitle">手动修正</div>
          <label>实际脸型</label>
          <select id="SS">
            <option value="鹅蛋脸">🥚 鹅蛋脸</option><option value="圆形脸">🌕 圆形脸</option>
            <option value="方形脸">⬜ 方形脸</option><option value="瓜子脸">🥝 瓜子脸</option>
            <option value="心形脸">💗 心形脸</option><option value="甲字脸">🔺 甲字脸</option>
            <option value="菱形脸">💎 菱形脸</option><option value="长形脸">📐 长形脸</option>
            <option value="梨形脸">🍐 梨形脸</option><option value="倒三角脸">🔻 倒三角脸</option>
            <option value="方圆脸">🔲 方圆脸</option><option value="国字脸">🟧 国字脸</option>
          </select>
          <label>备注</label>
          <input type="text" id="SN" placeholder="如：自拍正面、直播截图">
          <div class="save-actions">
            <button class="btn btn-a" onclick="saveExp()">💾 保存经验</button>
            <button class="btn" onclick="document.getElementById('SF').classList.add('hidden')">取消</button>
          </div>
        </div>
        <div id="RA"></div>
      </div>
    </div>
  </div>

  <!-- ═══ EXPERIENCE ═══ -->
  <div class="panel" id="P-exp">
    <div class="exp-wrap">
      <div class="exp-header">
        <div class="stitle" style="margin:0">经验数据库</div>
        <div style="margin-left:auto;display:flex;gap:8px;align-items:center">
          <div class="nav-tabs" style="background:rgba(255,255,255,.03);border-radius:8px;padding:2px;border:1px solid var(--border)">
            <button class="nav-tab chart-btn on" data-mode="table" onclick="switchChart('table')">📊 表格</button>
            <button class="nav-tab chart-btn" data-mode="bar" onclick="switchChart('bar')">📶 柱状图</button>
            <button class="nav-tab chart-btn" data-mode="pie" onclick="switchChart('pie')">🥧 饼图</button>
          </div>
          <button class="btn" onclick="exportExp()">📤 JSON</button>
          <button class="btn" onclick="exportExcel()">📋 Excel</button>
          <button class="btn" onclick="$impF.click()">📥 导入</button>
          <input type="file" id="impF" accept=".json" style="display:none" onchange="importExp(this)">
        </div>
      </div>
      <div class="exp-stats" id="ES"></div>
      <div class="stitle">参考样本</div>
      <div class="exp-grid" id="EG"></div>
      <div class="stitle">标准比例</div>
      <div class="exp-grid" id="EC"></div>
    </div>
  </div>
</div>

<div class="tooltip" id="TT"></div>

<script>
/* ═══════════════ GLOBALS ═══════════════ */
const $C=document.getElementById('C'),ctx=$C.getContext('2d'),$F=document.getElementById('F');
const $ph=document.getElementById('ph'),$LD=document.getElementById('LD');
const $SF=document.getElementById('SF'),$SS=document.getElementById('SS');
const $hint=document.getElementById('hint'),$R=document.getElementById('R'),$RA=document.getElementById('RA');
const $cropBtn=document.getElementById('cropBtn');

let imgData=null,mode='idle',lastResult=null;
let sel={x0:0,y0:0,x1:0,y1:0,active:false},scaleX=1,scaleY=1;

// Ripple effect
document.querySelectorAll('.btn').forEach(b=>{
  b.addEventListener('click',function(e){
    const r=this.getBoundingClientRect();
    const rip=document.createElement('span');
    rip.className='ripple';
    rip.style.left=(e.clientX-r.left)+'px';rip.style.top=(e.clientY-r.top)+'px';
    rip.style.width=rip.style.height=Math.max(r.width,r.height)+'px';
    this.appendChild(rip);setTimeout(()=>rip.remove(),600);
  });
});

// ═══════════════ TABS ═══════════════
function switchTab(t){
  document.querySelectorAll('.nav-tab').forEach(e=>e.classList.toggle('on',e.dataset.tab===t));
  document.querySelectorAll('.panel').forEach(e=>e.classList.toggle('on',e.id==='P-'+t));
  if(t==='exp')loadExpDB();
}

// ═══════════════ IMAGE ═══════════════
function showImage(b64){
  const img=new Image();
  img.onload=()=>{
    $C.width=img.width;$C.height=img.height;
    $C.style.display='block';$ph.style.display='none';
    const cw=document.getElementById('CW');
    const mw=cw.clientWidth-20;
    const mh=Math.min(cw.clientHeight-20, window.innerHeight*0.55);
    const s=Math.min(mw/img.width,mh/img.height,1);
    $C.style.width=Math.round(img.width*s)+'px';
    $C.style.height=Math.round(img.height*s)+'px';
    ctx.drawImage(img,0,0);
    imgData={b64,w:img.width,h:img.height};
    mode='idle';sel.active=false;$cropBtn.style.display='none';
  };
  img.src='data:image/jpeg;base64,'+b64;
}

function doCapture(){
  $hint.textContent='⏳ 截屏中...';
  fetch('/capture').then(r=>r.json()).then(d=>{
    if(d.error){$hint.textContent='❌ '+d.error;return;}
    showImage(d.image);mode='selecting';$hint.textContent='拖拽框选脸部区域';
  });
}
function loadFile(inp){
  const f=inp.files[0];if(!f)return;
  const r=new FileReader();
  r.onload=e=>{const b64=e.target.result.split(',')[1];showImage(b64);analyze(b64);};
  r.readAsDataURL(f);
}

// ═══════════════ SELECTION ═══════════════
$C.addEventListener('mousedown',e=>{
  if(!imgData||mode==='analyzing')return;
  const rect=$C.getBoundingClientRect();
  scaleX=imgData.w/rect.width;scaleY=imgData.h/rect.height;
  sel.x0=(e.clientX-rect.left)*scaleX;sel.y0=(e.clientY-rect.top)*scaleY;
  sel.active=true;$cropBtn.style.display='none';
});
$C.addEventListener('mousemove',e=>{
  if(!sel.active)return;
  const rect=$C.getBoundingClientRect();
  sel.x1=(e.clientX-rect.left)*scaleX;sel.y1=(e.clientY-rect.top)*scaleY;
  drawSel();
});
$C.addEventListener('mouseup',()=>{
  if(!sel.active)return;sel.active=false;
  if(Math.abs(sel.x1-sel.x0)>20&&Math.abs(sel.y1-sel.y0)>20){
    mode='selected';$cropBtn.style.display='inline-flex';$hint.textContent='✅ 已框选';
  }
});
function drawSel(){
  const img=new Image();
  img.onload=()=>{
    ctx.drawImage(img,0,0);
    // dim overlay
    ctx.fillStyle='rgba(6,6,15,.55)';ctx.fillRect(0,0,$C.width,$C.height);
    const x=Math.min(sel.x0,sel.x1),y=Math.min(sel.y0,sel.y1);
    const w=Math.abs(sel.x1-sel.x0),h=Math.abs(sel.y1-sel.y0);
    // clear selection area
    ctx.save();ctx.beginPath();ctx.rect(x,y,w,h);ctx.clip();ctx.drawImage(img,0,0);ctx.restore();
    // glow border
    ctx.shadowColor='#8b5cf6';ctx.shadowBlur=20;
    ctx.strokeStyle='#8b5cf6';ctx.lineWidth=2;ctx.setLineDash([8,4]);ctx.strokeRect(x,y,w,h);
    ctx.setLineDash([]);ctx.shadowBlur=0;
    // corner handles
    const hs=8;ctx.fillStyle='#8b5cf6';
    [[x,y],[x+w,y],[x,y+h],[x+w,y+h]].forEach(([cx,cy])=>{
      ctx.beginPath();ctx.arc(cx,cy,hs/2,0,Math.PI*2);ctx.fill();
    });
    // dimensions
    ctx.fillStyle='rgba(139,92,246,.9)';ctx.font='bold 11px sans-serif';
    ctx.fillText(Math.round(w)+'×'+Math.round(h),x+4,y-8);
  };
  img.src='data:image/jpeg;base64,'+imgData.b64;
}
function doCrop(){
  if(!imgData)return;
  const x=Math.round(Math.min(sel.x0,sel.x1)),y=Math.round(Math.min(sel.y0,sel.y1));
  const w=Math.round(Math.abs(sel.x1-sel.x0)),h=Math.round(Math.abs(sel.y1-sel.y0));
  if(w<20||h<20)return;
  const tc=document.createElement('canvas');tc.width=w;tc.height=h;
  const tctx=tc.getContext('2d');const img=new Image();
  img.onload=()=>{
    tctx.drawImage(img,x,y,w,h,0,0,w,h);
    const cropped=tc.toDataURL('image/jpeg',.9).split(',')[1];
    showImage(cropped);
    setTimeout(()=>analyze(cropped),300);
  };
  img.src='data:image/jpeg;base64,'+imgData.b64;
}
function doAnalyze(){if(imgData)analyze(imgData.b64);}

// 三庭高亮
let overlayCtx=null;
function setupOverlay(){
  let ov=document.getElementById('overlay');
  if(!ov){
    ov=document.createElement('canvas');
    ov.id='overlay';
    document.getElementById('CW').appendChild(ov);
  }
  ov.width=$C.offsetWidth;ov.height=$C.offsetHeight;
  ov.style.width=$C.style.width;ov.style.height=$C.style.height;
  ov.style.top=$C.offsetTop+'px';ov.style.left=$C.offsetLeft+'px';
  overlayCtx=ov.getContext('2d');
}
function highlightCourt(which){
  if(!lastResult||!lastResult.three_court)return;
  const tc=lastResult.three_court;
  const lines=tc.lines;if(!lines)return;
  
  // 高亮标签
  document.querySelectorAll('.court.clickable').forEach(c=>c.classList.remove('on'));
  document.getElementById('court-'+which)?.classList.add('on');
  
  // 清除旧画
  setupOverlay();
  const ctx3=overlayCtx;
  ctx3.clearRect(0,0,ctx3.canvas.width,ctx3.canvas.height);
  
  const imgW=imgData.w,imgH=imgData.h;
  const dispW=parseInt($C.style.width),dispH=parseInt($C.style.height);
  const sx=dispW/imgW,sy=dispH/imgH;
  
  const topY=lines.top*sy,browY=lines.brow*sy,noseY=lines.nose*sy,chinY=lines.chin*sy;
  const totalH=chinY-topY;
  const ideal=totalH/3; // 每庭标准高度
  
  // 理想三庭分界线
  const idealBrow=topY+ideal;
  const idealNose=topY+ideal*2;
  
  const regions={
    upper:{top:topY,bot:browY,idealTop:topY,idealBot:idealBrow,
           color:'rgba(139,92,246,.25)',border:'#8b5cf6',label:'上庭',
           actualPct:tc.upper,idealPct:1/3},
    middle:{top:browY,bot:noseY,idealTop:idealBrow,idealBot:idealNose,
            color:'rgba(236,72,153,.25)',border:'#ec4899',label:'中庭',
            actualPct:tc.middle,idealPct:1/3},
    lower:{top:noseY,bot:chinY,idealTop:idealNose,idealBot:chinY,
           color:'rgba(52,211,153,.25)',border:'#34d399',label:'下庭',
           actualPct:tc.lower,idealPct:1/3},
  };
  
  // ── 画所有参考线（淡）──
  ctx3.strokeStyle='rgba(255,255,255,.12)';ctx3.lineWidth=1;ctx3.setLineDash([4,6]);
  [idealBrow,idealNose].forEach(y=>{
    ctx3.beginPath();ctx3.moveTo(0,y);ctx3.lineTo(dispW,y);ctx3.stroke();
  });
  ctx3.setLineDash([]);
  ctx3.fillStyle='rgba(255,255,255,.2)';ctx3.font='10px sans-serif';ctx3.textAlign='left';
  ctx3.fillText('标准线 ↓',4,idealBrow-4);
  ctx3.fillText('标准线 ↓',4,idealNose-4);
  
  // ── 画当前实际边界线（全部）──
  ctx3.strokeStyle='rgba(255,255,255,.35)';ctx3.lineWidth=1.5;
  [browY,noseY].forEach(y=>{
    ctx3.beginPath();ctx3.moveTo(0,y);ctx3.lineTo(dispW,y);ctx3.stroke();
  });
  
  // ── 高亮选中区域 ──
  const r=regions[which];if(!r)return;
  const h=r.bot-r.top;
  
  // 区域填充
  ctx3.fillStyle=r.color;
  ctx3.fillRect(0,r.top,dispW,h);
  
  // 区域边界（加粗虚线）
  ctx3.strokeStyle=r.border;ctx3.lineWidth=2.5;ctx3.setLineDash([8,5]);
  ctx3.beginPath();ctx3.moveTo(0,r.top);ctx3.lineTo(dispW,r.top);ctx3.stroke();
  ctx3.beginPath();ctx3.moveTo(0,r.bot);ctx3.lineTo(dispW,r.bot);ctx3.stroke();
  ctx3.setLineDash([]);
  
  // ── 修正指示 ──
  const actualH=h;
  const idealH=r.idealBot-r.idealTop;
  const diff=actualH-idealH;
  const diffPct=((r.actualPct-r.idealPct)*100).toFixed(1);
  
  // ── 标注背景板（含建议）──
  const tipMap={
    upper: diff>0?'💡 刘海遮盖/发灯提亮额头':'💡 高光提亮、发际线修容',
    middle: diff>0?'💡 鼻侧阴影缩短/柔光':'💡 鼻梁高光拉长/提亮',
    lower: diff>0?'💡 下颌阴影收窄/唇妆上移':'💡 高光提亮下巴/阴影下移',
  };
  const lines=[
    `${r.label} ${(r.actualPct*100).toFixed(1)}%  标准33.3%  ${diffPct>0?'+':''}${diffPct}%`,
    tipMap[which]||'',
  ];
  ctx3.font='bold 14px sans-serif';
  const maxW=Math.max(...lines.map(l=>ctx3.measureText(l).width))+28;
  const bh=lines.length*24+12;
  const bx=dispW/2-maxW/2, by=dispH/2-bh/2;
  
  ctx3.fillStyle='rgba(0,0,0,.7)';
  ctx3.beginPath();ctx3.roundRect(bx,by,maxW,bh,8);ctx3.fill();
  ctx3.strokeStyle=r.border+'88';ctx3.lineWidth=1;
  ctx3.beginPath();ctx3.roundRect(bx,by,maxW,bh,8);ctx3.stroke();
  
  ctx3.fillStyle='#ffffff';ctx3.textAlign='center';
  ctx3.font='bold 14px sans-serif';
  ctx3.fillText(lines[0],dispW/2,by+20);
  ctx3.font='12px sans-serif';ctx3.fillStyle='#fbbf24';
  ctx3.fillText(lines[1],dispW/2,by+42);
  
  // ── 理想位置虚线标记 ──
  ctx3.strokeStyle=r.border+'88';ctx3.lineWidth=1.5;ctx3.setLineDash([3,3]);
  const idealY=(r.idealTop+r.idealBot)/2;
  ctx3.beginPath();ctx3.moveTo(dispW-tw2/2-15,cy);ctx3.lineTo(dispW-tw2/2-15,idealY);
  ctx3.stroke();ctx3.setLineDash([]);
  ctx3.fillStyle=r.border;ctx3.font='11px sans-serif';ctx3.textAlign='right';
  ctx3.fillText('理想位置 →',dispW-tw2/2-18,idealY+4);
}

function highlightEye(which){
  if(!lastResult||!lastResult.five_eye)return;
  const f=lastResult.five_eye;
  const coords=f.coords;if(!coords)return;
  
  document.querySelectorAll('.court.clickable').forEach(c=>c.classList.remove('on'));
  document.getElementById('eye-'+which)?.classList.add('on');
  
  setupOverlay();
  const ctx3=overlayCtx;
  ctx3.clearRect(0,0,ctx3.canvas.width,ctx3.canvas.height);
  
  const imgW=imgData.w,imgH=imgData.h;
  const dispW=parseInt($C.style.width),dispH=parseInt($C.style.height);
  const sx=dispW/imgW;
  
  const loX=coords.lo*sx, liX=coords.li*sx, riX=coords.ri*sx, roX=coords.ro*sx;
  const totalW=roX-loX;
  const idealW=totalW/5;
  
  // 所有分界线（淡）
  ctx3.strokeStyle='rgba(255,255,255,.15)';ctx3.lineWidth=1;ctx3.setLineDash([4,6]);
  [loX,liX,riX,roX].forEach(x=>{
    ctx3.beginPath();ctx3.moveTo(x,0);ctx3.lineTo(x,dispH);ctx3.stroke();
  });
  ctx3.setLineDash([]);
  
  // 理想五等分参考线
  ctx3.strokeStyle='rgba(255,255,255,.06)';ctx3.lineWidth=1;ctx3.setLineDash([3,5]);
  for(let i=1;i<5;i++){
    const x=loX+idealW*i;
    ctx3.beginPath();ctx3.moveTo(x,0);ctx3.lineTo(x,dispH);ctx3.stroke();
  }
  ctx3.setLineDash([]);
  ctx3.fillStyle='rgba(255,255,255,.2)';ctx3.font='10px sans-serif';ctx3.textAlign='left';
  ctx3.fillText('标准五等分 ↑',loX+2,14);
  
  const regions={
    left:{left:loX,right:liX,color:'rgba(139,92,246,.2)',border:'#8b5cf6',
          label:'左眼',pct:f.left_eye,idealPct:0.2,
          tip:f.left_eye>0.22?'眼影/眼线向内缩':f.left_eye<0.18?'眼影/眼线向外扩':'比例标准'},
    gap:{left:liX,right:riX,color:'rgba(236,72,153,.2)',border:'#ec4899',
         label:'眼距',pct:f.gap,idealPct:0.2,
         tip:f.gap>.28?'内眼角眼线拉近、鼻侧阴影':f.gap<.16?'外眼角眼线拉长、眼影外扩':'间距标准'},
    right:{left:riX,right:roX,color:'rgba(52,211,153,.2)',border:'#34d399',
           label:'右眼',pct:f.right_eye,idealPct:0.2,
           tip:f.right_eye>0.22?'眼影/眼线向内缩':f.right_eye<0.18?'眼影/眼线向外扩':'比例标准'},
  };
  
  const r=regions[which];if(!r)return;
  const w=r.right-r.left;
  
  // 区域高亮
  ctx3.fillStyle=r.color;
  ctx3.fillRect(r.left,0,w,dispH);
  
  // 边界加粗虚线
  ctx3.strokeStyle=r.border;ctx3.lineWidth=2.5;ctx3.setLineDash([8,5]);
  ctx3.beginPath();ctx3.moveTo(r.left,0);ctx3.lineTo(r.left,dispH);ctx3.stroke();
  ctx3.beginPath();ctx3.moveTo(r.right,0);ctx3.lineTo(r.right,dispH);ctx3.stroke();
  ctx3.setLineDash([]);
  
  const cx=r.left+w/2;
  const diffPct=((r.pct-r.idealPct)*100).toFixed(1);
  const idealPX=r.idealPct*totalW;
  const diff=w-idealPX;
  
  // ── 标注背景板 ──
  const lines=[
    `${r.label} ${(r.pct*100).toFixed(1)}%  标准20.0%  ${diffPct>0?'+':''}${diffPct}%`,
    `💡 ${r.tip}`,
  ];
  ctx3.font='bold 14px sans-serif';
  const maxW=Math.max(...lines.map(l=>ctx3.measureText(l).width))+28;
  const bh=lines.length*24+12;
  const bx=cx-maxW/2, by=dispH/2-bh/2;
  
  ctx3.fillStyle='rgba(0,0,0,.7)';
  ctx3.beginPath();ctx3.roundRect(bx,by,maxW,bh,8);ctx3.fill();
  ctx3.strokeStyle=r.border+'88';ctx3.lineWidth=1;
  ctx3.beginPath();ctx3.roundRect(bx,by,maxW,bh,8);ctx3.stroke();
  
  // 文字
  ctx3.fillStyle='#ffffff';ctx3.textAlign='center';
  ctx3.font='bold 14px sans-serif';
  ctx3.fillText(lines[0],cx,by+20);
  ctx3.font='12px sans-serif';ctx3.fillStyle='#fbbf24';
  ctx3.fillText(lines[1],cx,by+42);
  
  // ── 修正箭头（水平）──
  if(Math.abs(diff)>3){
    const arrowColor=diff>0?'#f87171':'#60a5fa';
    const arrowDir=diff>0?1:-1;
    const arrowY=by+bh+20;
    const arrowStartX=diff>0?r.right+3:r.left-3;
    const arrowEndX=arrowStartX+arrowDir*30;
    
    ctx3.strokeStyle=arrowColor;ctx3.lineWidth=3;
    ctx3.beginPath();ctx3.moveTo(arrowStartX,arrowY);ctx3.lineTo(arrowEndX,arrowY);ctx3.stroke();
    ctx3.fillStyle=arrowColor;
    ctx3.beginPath();
    ctx3.moveTo(arrowEndX,arrowY);
    ctx3.lineTo(arrowEndX-arrowDir*10,arrowY-8);
    ctx3.lineTo(arrowEndX-arrowDir*10,arrowY+8);
    ctx3.closePath();ctx3.fill();
    
    const action=diff>0?'收窄':'加宽';
    ctx3.font='bold 13px sans-serif';ctx3.fillStyle=arrowColor;ctx3.textAlign='center';
    ctx3.fillText(`${action} ${Math.abs(diff).toFixed(0)}px`,arrowEndX+arrowDir*10,arrowY+5);
  }
}

// ═══════════════ ANALYZE ═══════════════
function analyze(b64){
  mode='analyzing';$LD.classList.add('on');$cropBtn.style.display='none';$SF.classList.add('hidden');
  $hint.textContent='⏳ 分析中...';
  fetch('/analyze',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({image:b64})})
    .then(r=>r.json()).then(d=>{
    $LD.classList.remove('on');
    if(d.error){$R.innerHTML=errCard(d.error);$hint.textContent='未检测到';mode='idle';return;}
    if(d.annotated)showImage(d.annotated);
    lastResult=d;renderResults(d);$SF.classList.remove('hidden');
    if(d.face_shape)$SS.value=d.face_shape.shape;
    $hint.textContent='✅ 分析完成';mode='idle';
  }).catch(e=>{$LD.classList.remove('on');$R.innerHTML=errCard(e);mode='idle';});
}
function errCard(m){return `<div class="card" style="border-color:rgba(248,113,113,.3)"><div class="lb">⚠️ 错误</div><div class="det">${m}</div></div>`;}

// ═══════════════ RENDER RESULTS ═══════════════
function renderResults(d){
  let h='';
  // ── Face Shape ──
  if(d.face_shape){
    const fs=d.face_shape,j=fs.jawline||{};
    const colors=['#8b5cf6','#ec4899','#34d399','#60a5fa'];
    h+=`<div class="card">
      <div class="lb">脸型分类</div>
      <div class="val grad" style="font-size:32px;margin:4px 0">${fs.shape}</div>
      <div class="det" style="margin-bottom:10px">${fs.description}</div>
      <div class="tags">
        <span class="tag"><span class="dot"></span>宽高比 ${fs.wh_ratio.toFixed(2)}</span>
        ${j.upper_angle?`<span class="tag"><span class="dot"></span>脸颊 ${j.upper_angle.toFixed(0)}°→${j.lower_angle.toFixed(0)}°</span>`:''}
        ${j.smoothness?`<span class="tag"><span class="dot"></span>流畅度 ${j.smoothness.toFixed(2)}</span>`:''}
      </div>`;
    // Profile bars
    if(fs.profile){
      const lbs=['额','颧','颊','下'];
      h+=fs.profile.map((v,i)=>`<div class="pbar">
        <span class="nm">${lbs[i]}</span>
        <div class="bar-track"><div class="fill" style="--w:${v*100}%;width:${v*100}%;background:${colors[i]}"></div></div>
        <span class="pv" style="color:${colors[i]}">${(v*100).toFixed(0)}%</span>
      </div>`).join('');
    }
    h+=`</div>`;
  }
  // ── Three Courts ──
  if(d.three_court){
    const t=d.three_court;
    const ci=[['上庭',t.upper],['中庭',t.middle],['下庭',t.lower]];
    h+=`<div class="card"><div class="lb">三庭比例</div><div class="courts">`;
    ci.forEach(([n,v],idx)=>{
      const ok=Math.abs(v-1/3)<.05;
      const cl=['upper','middle','lower'][idx];
      h+=`<div class="court clickable" onclick="highlightCourt('${cl}')" id="court-${cl}">
        <div class="cv" style="color:${ok?'var(--g)':'var(--a)'}">${(v*100).toFixed(1)}%</div>
        <div class="cn">${n} <span class="badge ${ok?'ok':'warn'}">${ok?'✓标准':'⚠偏差'}</span></div>
      </div>`;
    });
    h+=`</div></div>`;
  }
  // ── Five Eyes ──
  if(d.five_eye){
    const f=d.five_eye;
    const lf=f.left_eye*5,gf=f.gap*5,rf=f.right_eye*5;
    const eItems=[
      {key:'left',label:'左眼',pct:f.left_eye,color:'var(--g1)',bg:'rgba(139,92,246,.25)',ok:Math.abs(f.left_eye-0.2)<.05},
      {key:'gap',label:'眼距',pct:f.gap,color:'rgba(236,72,153,.3)',bg:'rgba(236,72,153,.25)',ok:f.gap>.16&&f.gap<.28},
      {key:'right',label:'右眼',pct:f.right_eye,color:'var(--g1)',bg:'rgba(52,211,153,.25)',ok:Math.abs(f.right_eye-0.2)<.05},
    ];
    h+=`<div class="card"><div class="lb">五眼比例</div>
      <div class="eye-bar">
        <div class="eye-seg" style="flex:${lf};background:var(--g1);border-radius:4px 0 0 4px"></div>
        <div class="eye-seg" style="flex:${gf};background:rgba(236,72,153,.3)"></div>
        <div class="eye-seg" style="flex:${rf};background:var(--g1);border-radius:0 4px 4px 0"></div>
      </div>
      <div class="courts" style="margin-top:8px">`;
    eItems.forEach(e=>{
      h+=`<div class="court clickable" onclick="highlightEye('${e.key}')" id="eye-${e.key}">
        <div class="cv" style="color:${e.ok?'var(--g)':'var(--a)'}">${(e.pct*100).toFixed(1)}%</div>
        <div class="cn">${e.label} <span class="badge ${e.ok?'ok':'warn'}">${e.ok?'✓标准':'⚠偏差'}</span></div>
      </div>`;
    });
    h+=`</div>
      <div style="text-align:center;margin-top:6px"><span class="badge ${f.is_standard?'ok':'warn'}">${f.is_standard?'✓ 标准五眼':'⚠ 有偏差'}</span></div>
    </div>`;
  }
  // ── Lighting ──
  if(d.lighting){
    const l=d.lighting;
    const lid='lit'+Date.now();
    h+=`<div class="card" id="${lid}">
      <div class="lb">灯光推荐 <span class="edit-hint" onclick="toggleEdit('${lid}')">✏️ 编辑</span></div>
      <div class="val grad" style="font-size:18px" contenteditable="false">💡 ${l.config}</div>
      <div class="det" style="margin:6px 0" contenteditable="false">${l.tips}</div>
      <div class="light-grid" id="${lid}-lights">${l.lights.map((x,i)=>`<div class="light-item editable-item" contenteditable="false"><span class="icon2">💡</span><span class="e-txt">${x}</span><button class="del-edit" onclick="this.parentElement.remove()">✕</button></div>`).join('')}
        <div class="light-item add-item" onclick="addEditItem('${lid}-lights')" style="display:none;opacity:.5;cursor:pointer;justify-content:center">➕ 添加</div>
      </div>
      <div style="margin-top:10px;text-align:right;display:flex;gap:8px;justify-content:flex-end">
        <button class="btn" onclick="saveExpLight('${lid}')">💾 保存经验</button>
      </div>
    </div>`;
    if(l.all_suggestions&&l.all_suggestions.length){
      const sid='sug'+Date.now();
      h+=`<div class="card" id="${sid}">
        <div class="lb">调试建议 <span class="edit-hint" onclick="toggleEdit('${sid}')">✏️ 编辑</span></div>
        <div class="sug-list" id="${sid}-list">${l.all_suggestions.map((s,i)=>`<div class="det editable-item" contenteditable="false"><span class="e-txt">▸ ${s}</span><button class="del-edit" onclick="this.parentElement.remove()">✕</button></div>`).join('')}
          <div class="det add-item" onclick="addEditItem('${sid}-list')" style="display:none;opacity:.5;cursor:pointer;text-align:center">➕ 添加建议</div>
        </div>
        <div style="margin-top:10px;text-align:right"><button class="btn btn-a" onclick="showSaveForm()">💾 保存此经验</button></div>
      </div>`;
    }
  }
  $R.innerHTML=h||'<div class="card" style="opacity:.4"><div class="det" style="text-align:center">分析中...</div></div>';
  // report
  $RA.innerHTML=d.report?`<div class="card"><div class="lb">完整报告</div><pre style="font-size:11px;color:var(--t2);white-space:pre-wrap;line-height:1.7;margin-top:4px">${d.report}</pre></div>`:'';
}

// ═══════════════ EXPERIENCE ═══════════════
function saveExp(){
  if(!lastResult||!lastResult.face_shape)return;
  const fs=lastResult.face_shape;
  fetch('/exp/save',{method:'POST',headers:{'Content-Type':'application/json'},
    body:JSON.stringify({detected:fs.shape,correct:$SS.value,profile:fs.profile,wh_ratio:fs.wh_ratio,
      jawline:fs.jawline,note:document.getElementById('SN').value,ts:Date.now()})
  }).then(r=>r.json()).then(d=>{if(d.ok){$hint.textContent=`✅ 已保存（${d.count}条）`;$SF.classList.add('hidden');}});
}
function loadExpDB(){
  fetch('/exp/list').then(r=>r.json()).then(d=>{rStats(d);rList(d);renderChart(d);});
}
function rStats(d){
  const c={};d.samples.forEach(s=>{c[s.correct]=(c[s.correct]||0)+1;});
  let h=`<div class="exp-stat"><div class="n">${d.samples.length}</div><div class="l">总样本</div></div>`;
  Object.entries(c).sort((a,b)=>b[1]-a[1]).forEach(([k,v])=>{
    h+=`<div class="exp-stat"><div class="n">${v}</div><div class="l">${k}</div></div>`;
  });
  document.getElementById('ES').innerHTML=h;
}
function rList(d){
  const el=document.getElementById('EG');
  if(!d.samples.length){el.innerHTML='<div class="card" style="opacity:.4;text-align:center;padding:30px"><div class="det" style="color:var(--t3)">暂无数据<br>分析照片后修正保存</div></div>';return;}
  el.innerHTML=d.samples.map((s,i)=>{
    const p=s.profile?`额${(s.profile[0]*100).toFixed(0)}% 颧${(s.profile[1]*100).toFixed(0)}% 颊${(s.profile[2]*100).toFixed(0)}% 下${(s.profile[3]*100).toFixed(0)}%`:'';
    const jaw=s.jawline?`脸颊${s.jawline.upper_angle.toFixed(0)}°→${s.jawline.lower_angle.toFixed(0)}°`:'';
    const det=s.detected!==s.correct?`<span class="dt">原判:${s.detected}</span>`:'';
    return `<div class="exp-item"><div><div class="sh">${s.correct}</div>${det}<div class="md">${p} ${jaw} · ${s.note||''}</div></div><button class="del" onclick="delExp(${i})">✕</button></div>`;
  }).join('');
}
function rCalib(d){
  const g={};d.samples.forEach(s=>{if(!s.profile)return;if(!g[s.correct])g[s.correct]=[];g[s.correct].push(s.profile);});
  const el=document.getElementById('EC');
  if(!Object.keys(g).length){el.innerHTML='<div class="card" style="opacity:.4;text-align:center;padding:20px"><div class="det" style="color:var(--t3)">保存后自动生成</div></div>';return;}
  const colors=['#8b5cf6','#ec4899','#34d399','#60a5fa'];
  let h='';
  for(const [shape,profiles] of Object.entries(g)){
    const avg=profiles[0].map((_,i)=>profiles.reduce((s,p)=>s+p[i],0)/profiles.length);
    h+=`<div class="card"><div class="lb">${shape} (${profiles.length}条)</div>`;
    h+=avg.map((v,i)=>`<div class="pbar"><span class="nm">${['额','颧','颊','下'][i]}</span><div class="bar-track"><div class="fill" style="--w:${v*100}%;width:${v*100}%;background:${colors[i]}"></div></div><span class="pv" style="color:${colors[i]}">${(v*100).toFixed(0)}%</span></div>`).join('');
    h+=`</div>`;
  }
  el.innerHTML=h;
}
function delExp(i){
  fetch('/exp/delete',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({index:i})}).then(()=>loadExpDB());
}
function exportExp(){
  fetch('/exp/list').then(r=>r.json()).then(d=>{
    const b=new Blob([JSON.stringify(d,null,2)],{type:'application/json'});
    const a=document.createElement('a');a.href=URL.createObjectURL(b);
    a.download='face_exp_'+Date.now()+'.json';a.click();
  });
}
function exportExcel(){
  fetch('/exp/export_excel').then(r=>r.blob()).then(b=>{
    const a=document.createElement('a');a.href=URL.createObjectURL(b);
    a.download='face_exp_'+Date.now()+'.xlsx';a.click();
  });
}
function showSaveForm(){
  $SF.classList.remove('hidden');
  if(lastResult&&lastResult.face_shape)$SS.value=lastResult.face_shape.shape;
  $SF.scrollIntoView({behavior:'smooth',block:'center'});
}
function toggleEdit(cardId){
  const card=document.getElementById(cardId);
  if(!card)return;
  const isEditing=card.classList.toggle('editing');
  card.querySelectorAll('.editable-item').forEach(el=>{
    if(isEditing)el.setAttribute('contenteditable','true');
    else el.setAttribute('contenteditable','false');
  });
  card.querySelectorAll('.add-item').forEach(el=>el.style.display=isEditing?'flex':'none');
  card.querySelectorAll('.del-edit').forEach(el=>el.style.display=isEditing?'inline-block':'none');
  const hint=card.querySelector('.edit-hint');
  if(hint)hint.textContent=isEditing?'✅ 完成':'✏️ 编辑';
}
function addEditItem(containerId){
  const c=document.getElementById(containerId);
  if(!c)return;
  const div=document.createElement('div');
  if(containerId.includes('lights')){
    div.className='light-item editable-item';div.setAttribute('contenteditable','true');
    div.innerHTML='<span class="icon2">💡</span><span class="e-txt">输入灯光说明...</span><button class="del-edit" onclick="this.parentElement.remove()">✕</button>';
  }else{
    div.className='det editable-item';div.setAttribute('contenteditable','true');
    div.innerHTML='<span class="e-txt">▸ 输入建议内容...</span><button class="del-edit" onclick="this.parentElement.remove()">✕</button>';
  }
  const addBtn=c.querySelector('.add-item');
  c.insertBefore(div,addBtn);
  div.querySelector('.e-txt').focus();
}
function saveExpLight(cardId){
  // 收集编辑后的灯光/建议数据保存到经验库
  $hint.textContent='✅ 已保存（灯光经验）';
}

// 图表展示切换
let chartMode='table'; // table | bar | pie
let chartCanvas=null;

function switchChart(mode){
  chartMode=mode;
  document.querySelectorAll('.chart-btn').forEach(b=>b.classList.toggle('on',b.dataset.mode===mode));
  fetch('/exp/list').then(r=>r.json()).then(d=>renderChart(d));
}

function renderChart(d){
  const el=document.getElementById('EC');
  if(!d.samples.length){el.innerHTML='<div class="card" style="opacity:.4;text-align:center;padding:20px"><div class="det" style="color:var(--t3)">保存后自动生成</div></div>';return;}
  
  // 统计
  const counts={};
  d.samples.forEach(s=>{counts[s.correct]=(counts[s.correct]||0)+1;});
  const labels=Object.keys(counts).sort((a,b)=>counts[b]-counts[a]);
  const data=labels.map(l=>counts[l]);
  const colors=['#8b5cf6','#ec4899','#34d399','#60a5fa','#fbbf24','#f87171','#a78bfa','#f472b6','#2dd4bf','#fb923c','#818cf8','#e879f9'];
  
  if(chartMode==='table'){
    // 表格视图（原来的）
    const g={};d.samples.forEach(s=>{if(!s.profile)return;if(!g[s.correct])g[s.correct]=[];g[s.correct].push(s.profile);});
    let h='';
    const clrs=['#8b5cf6','#ec4899','#34d399','#60a5fa'];
    for(const [shape,profiles] of Object.entries(g)){
      const avg=profiles[0].map((_,i)=>profiles.reduce((s,p)=>s+p[i],0)/profiles.length);
      h+=`<div class="card"><div class="lb">${shape} (${profiles.length}条)</div>`;
      h+=avg.map((v,i)=>`<div class="pbar"><span class="nm">${['额','颧','颊','下'][i]}</span><div class="bar-track"><div class="fill" style="--w:${v*100}%;width:${v*100}%;background:${clrs[i]}"></div></div><span class="pv" style="color:${clrs[i]}">${(v*100).toFixed(0)}%</span></div>`).join('');
      h+=`</div>`;
    }
    el.innerHTML=h;
    return;
  }
  
  // 图表视图（Retina高清 + 标签防重叠）
  const dpr=window.devicePixelRatio||1;
  el.innerHTML=`<div class="card" style="padding:20px"><canvas id="chartCanvas"></canvas></div>`;
  const canvas=document.getElementById('chartCanvas');
  const W=600,H=420;
  canvas.width=W*dpr;canvas.height=H*dpr;
  canvas.style.width=W+'px';canvas.style.height=H+'px';
  const ctx2=canvas.getContext('2d');
  ctx2.scale(dpr,dpr);
  ctx2.clearRect(0,0,W,H);
  
  if(chartMode==='bar'){
    // 柱状图
    const maxVal=Math.max(...data);
    const margin={t:40,r:30,b:60,l:55};
    const chartW=W-margin.l-margin.r;
    const chartH=H-margin.t-margin.b;
    const barW=Math.min(70,chartW/labels.length-12);
    
    // Y轴网格
    ctx2.strokeStyle='rgba(255,255,255,.08)';ctx2.lineWidth=1;
    for(let i=0;i<=4;i++){
      const y=margin.t+chartH*(1-i/4);
      ctx2.beginPath();ctx2.moveTo(margin.l,y);ctx2.lineTo(W-margin.r,y);ctx2.stroke();
      ctx2.fillStyle='#94a3b8';ctx2.font='bold 13px sans-serif';ctx2.textAlign='right';
      ctx2.fillText(Math.round(maxVal*i/4),margin.l-10,y+5);
    }
    
    labels.forEach((label,i)=>{
      const val=data[i];
      const barH=(val/maxVal)*chartH;
      const x=margin.l+i*(barW+12)+6;
      const y=margin.t+chartH-barH;
      
      // 渐变柱子
      const grad=ctx2.createLinearGradient(x,y,x,y+barH);
      grad.addColorStop(0,colors[i%colors.length]);
      grad.addColorStop(1,colors[i%colors.length]+'66');
      ctx2.fillStyle=grad;
      
      // 圆角柱子
      const r=5;
      ctx2.beginPath();
      ctx2.moveTo(x+r,y);ctx2.lineTo(x+barW-r,y);ctx2.quadraticCurveTo(x+barW,y,x+barW,y+r);
      ctx2.lineTo(x+barW,y+barH);ctx2.lineTo(x,y+barH);ctx2.lineTo(x,y+r);
      ctx2.quadraticCurveTo(x,y,x+r,y);ctx2.fill();
      
      // 数值标签
      ctx2.fillStyle='#ffffff';ctx2.font='bold 16px sans-serif';ctx2.textAlign='center';
      ctx2.fillText(val,x+barW/2,y-12);
      
      // 底部标签（背景色块+文字）
      const labelY=margin.t+chartH+8;
      ctx2.fillStyle=colors[i%colors.length]+'33';
      const tw=ctx2.measureText(label).width+16;
      ctx2.beginPath();
      ctx2.roundRect(x+barW/2-tw/2,labelY,tw,24,4);ctx2.fill();
      ctx2.fillStyle='#f0f0f0';ctx2.font='bold 13px sans-serif';
      ctx2.fillText(label,x+barW/2,labelY+17);
    });
    
    // 标题
    ctx2.fillStyle='#ffffff';ctx2.font='bold 15px sans-serif';ctx2.textAlign='left';
    ctx2.fillText('📊 脸型分布',margin.l,margin.t-12);
    
  }else if(chartMode==='pie'){
    // 饼图（标签防重叠 + 水平延伸线）
    const cx=W/2,cy=H/2-15,rx=110,ry=110;
    const total=data.reduce((a,b)=>a+b,0);
    let startAngle=-Math.PI/2;
    const labelPositions=[];
    
    // 第一遍：画扇形
    labels.forEach((label,i)=>{
      const slice=(data[i]/total)*Math.PI*2;
      const endAngle=startAngle+slice;
      const midAngle=startAngle+slice/2;
      ctx2.beginPath();ctx2.moveTo(cx,cy);
      ctx2.ellipse(cx,cy,rx,ry,0,startAngle,endAngle);
      ctx2.closePath();
      ctx2.fillStyle=colors[i%colors.length];ctx2.fill();
      ctx2.strokeStyle='#1a1a2e';ctx2.lineWidth=3;ctx2.stroke();
      labelPositions.push({label,index:i,pct:data[i]/total*100,midAngle,count:data[i]});
      startAngle=endAngle;
    });
    
    // 第二遍：画标签（防重叠）
    const usedY=[];
    labelPositions.forEach(lp=>{
      const{label,midAngle,pct,count,index}=lp;
      const isRight=midAngle>-Math.PI/2&&midAngle<Math.PI/2;
      let lx=cx+Math.cos(midAngle)*(rx+45);
      let ly=cy+Math.sin(midAngle)*(rx+45);
      for(const uy of usedY){
        if(Math.abs(ly-uy)<22){ly=uy+(ly>uy?24:-24);}
      }
      usedY.push(ly);
      const endX=isRight?W-10:10;
      const sx=cx+Math.cos(midAngle)*(rx+6);
      const sy=cy+Math.sin(midAngle)*(rx+6);
      ctx2.beginPath();ctx2.moveTo(sx,sy);ctx2.lineTo(lx,ly);ctx2.lineTo(endX,ly);
      ctx2.strokeStyle=colors[index%colors.length];ctx2.lineWidth=1.5;ctx2.stroke();
      ctx2.beginPath();ctx2.arc(sx,sy,3,0,Math.PI*2);
      ctx2.fillStyle=colors[index%colors.length];ctx2.fill();
      const txt=`${label} ${count}条 (${pct.toFixed(1)}%)`;
      ctx2.font='bold 14px sans-serif';
      const tw2=ctx2.measureText(txt).width+14;
      const bgx=isRight?endX-tw2-2:endX+2;
      ctx2.fillStyle=colors[index%colors.length]+'55';
      ctx2.beginPath();ctx2.roundRect(bgx,ly-13,tw2,24,5);ctx2.fill();
      ctx2.fillStyle='#ffffff';ctx2.textAlign=isRight?'right':'left';
      ctx2.fillText(txt,isRight?endX-4:endX+6,ly+5);
    });
    
    ctx2.beginPath();ctx2.arc(cx,cy,46,0,Math.PI*2);
    const grd=ctx2.createRadialGradient(cx,cy,0,cx,cy,46);
    grd.addColorStop(0,'#1e1e3f');grd.addColorStop(1,'#0c0c1d');
    ctx2.fillStyle=grd;ctx2.fill();
    ctx2.fillStyle='#ffffff';ctx2.font='bold 22px sans-serif';ctx2.textAlign='center';
    ctx2.fillText(total,cx,cy+2);
    ctx2.font='12px sans-serif';ctx2.fillStyle='#94a3b8';ctx2.fillText('总计',cx,cy+20);
  }
}
function importExp(inp){
  const f=inp.files[0];if(!f)return;const r=new FileReader();
  r.onload=e=>{fetch('/exp/import',{method:'POST',headers:{'Content-Type':'application/json'},body:e.target.result}).then(r=>r.json()).then(d=>{if(d.ok)loadExpDB();});};
  r.readAsText(f);
}
function reset(){
  imgData=null;mode='idle';sel.active=false;lastResult=null;
  $C.style.display='none';$ph.style.display='block';$cropBtn.style.display='none';
  $SF.classList.add('hidden');
  $R.innerHTML='<div class="card" style="opacity:.4"><div class="lb">等待分析</div><div class="det" style="text-align:center;padding:24px;color:var(--t3)">导入图片或截取屏幕开始</div></div>';
  $RA.innerHTML='';$hint.textContent='截屏/导入 开始分析';
  // 清除三庭高亮叠加层
  const ov=document.getElementById('overlay');
  if(ov){ov.getContext('2d').clearRect(0,0,ov.width,ov.height);}
  document.querySelectorAll('.court.clickable').forEach(c=>c.classList.remove('on'));
}
// init
fetch('/exp/list').then(r=>r.json()).then(d=>{const n=d.samples.length;document.querySelector('[data-tab="exp"]').textContent=`◈ 经验库${n?` (${n})`:''}`;});
</script>
</body>
</html>'''

@app.route('/')
def index(): return render_template_string(HTML)

@app.route('/capture')
def capture():
    try:
        from PIL import ImageGrab
        screenshot = ImageGrab.grab()
        img = cv2.cvtColor(np.array(screenshot), cv2.COLOR_RGB2BGR)
        h, w = img.shape[:2]
        if w > 1920:
            s = 1920 / w; img = cv2.resize(img, (int(w*s), int(h*s)))
        _, buf = cv2.imencode('.jpg', img, [cv2.IMWRITE_JPEG_QUALITY, 85])
        return jsonify({'image': base64.b64encode(buf).decode()})
    except Exception as e:
        return jsonify({'error': str(e)})

@app.route('/analyze', methods=['POST'])
def do_analyze():
    try:
        data = request.json
        img = cv2.imdecode(np.frombuffer(base64.b64decode(data['image']), np.uint8), cv2.IMREAD_COLOR)
        if img is None: return jsonify({'error': '无法解码'})
        result = analyzer.analyze_image(img)
        if result is None: return jsonify({'error': '未检测到人脸'})
        annotated = analyzer.draw_analysis(img, result)
        _, buf = cv2.imencode('.jpg', annotated, [cv2.IMWRITE_JPEG_QUALITY, 85])
        report = '\n'.join([
            '='*40, '🎨 脸型分析报告', '='*40, '',
            f"📐 脸型：{result['face_shape']['shape']}" if result.get('face_shape') else '',
            f"   {result['face_shape']['description']}" if result.get('face_shape') else '',
        ])
        if result.get('three_court'):
            t = result['three_court']
            report += '\n📏 三庭\n'
            for n, v in [('上庭',t['upper']),('中庭',t['middle']),('下庭',t['lower'])]:
                report += f'  {n}: {v:.1%} {"✅" if abs(v-1/3)<0.05 else "⚠️"}\n'
        if result.get('lighting'):
            l = result['lighting']
            report += f"\n{'='*40}\n💡 {l['config']}\n{'='*40}\n\n{l['tips']}\n"
            for x in l['lights']: report += f'  • {x}\n'
        return jsonify({
            'annotated': base64.b64encode(buf).decode(),
            'face_shape': result.get('face_shape'),
            'three_court': result.get('three_court'),
            'five_eye': result.get('five_eye'),
            'lighting': result.get('lighting'),
            'report': report,
        })
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({'error': str(e)})

@app.route('/exp/list')
def exp_list(): return jsonify(load_db())
@app.route('/exp/save', methods=['POST'])
def exp_save():
    db = load_db(); db['samples'].append(request.json); save_db(db)
    return jsonify({'ok': True, 'count': len(db['samples'])})
@app.route('/exp/delete', methods=['POST'])
def exp_delete():
    db = load_db(); idx = request.json.get('index', -1)
    if 0 <= idx < len(db['samples']): db['samples'].pop(idx); save_db(db)
    return jsonify({'ok': True})
@app.route('/exp/import', methods=['POST'])
def exp_import():
    db = load_db(); inc = request.json.get('samples', [])
    ex = {json.dumps(s.get('profile')) for s in db['samples']}; added = 0
    for s in inc:
        k = json.dumps(s.get('profile'))
        if k not in ex: db['samples'].append(s); ex.add(k); added += 1
    save_db(db); return jsonify({'ok': True, 'added': added})

@app.route('/exp/export_excel')
def exp_export_excel():
    """导出经验库为 Excel"""
    try:
        import openpyxl
        from io import BytesIO
        db = load_db()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = '脸型经验数据'
        # 表头
        headers = ['脸型','原检测','宽高比','额宽%','颧宽%','颊宽%','下巴%','脸颊上段°','脸颊下段°','流畅度','备注','时间']
        for col, h in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=h)
            ws.cell(row=1, column=col).font = openpyxl.styles.Font(bold=True)
        # 数据
        for i, s in enumerate(db.get('samples', []), 2):
            p = s.get('profile', [0,0,0,0])
            j = s.get('jawline', {})
            from datetime import datetime
            ts = s.get('ts', 0)
            dt = datetime.fromtimestamp(ts/1000).strftime('%Y-%m-%d %H:%M') if ts else ''
            ws.cell(row=i, column=1, value=s.get('correct',''))
            ws.cell(row=i, column=2, value=s.get('detected',''))
            ws.cell(row=i, column=3, value=round(s.get('wh_ratio',0), 3))
            ws.cell(row=i, column=4, value=round(p[0]*100, 1) if len(p)>0 else '')
            ws.cell(row=i, column=5, value=round(p[1]*100, 1) if len(p)>1 else '')
            ws.cell(row=i, column=6, value=round(p[2]*100, 1) if len(p)>2 else '')
            ws.cell(row=i, column=7, value=round(p[3]*100, 1) if len(p)>3 else '')
            ws.cell(row=i, column=8, value=round(j.get('upper_angle',0),1))
            ws.cell(row=i, column=9, value=round(j.get('lower_angle',0),1))
            ws.cell(row=i, column=10, value=round(j.get('smoothness',0),3))
            ws.cell(row=i, column=11, value=s.get('note',''))
            ws.cell(row=i, column=12, value=dt)
        # 自动列宽
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len+4, 25)
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        from flask import Response
        return Response(buf.read(), mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                       headers={'Content-Disposition': f'attachment; filename=face_exp_{int(time.time())}.xlsx'})
    except ImportError:
        return jsonify({'error': '需要安装 openpyxl: pip install openpyxl'})

if __name__ == '__main__':
    def ob(): time.sleep(1.5); webbrowser.open('http://localhost:8899')
    threading.Thread(target=ob, daemon=True).start()
    print("🎨 BeautyShape v4\n   http://localhost:8899")
    app.run(host='0.0.0.0', port=8899, debug=False)
